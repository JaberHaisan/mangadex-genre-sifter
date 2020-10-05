# excel_link_opener.py - Opens all links in a particular column in a chosen
# excel document.

import openpyxl
import webbrowser
import time
import os

folder_name = "Output"
msg = "Please run mangadex_sifter.py first and then this script."

def link_opener(file_path, col=3):
	"""Opens all links in the given column in an
	excel document."""	
	wb = openpyxl.load_workbook(file_path)
	ws = wb.active
	links = []
	for i in range(2, ws.max_row + 1):
		links.append(ws.cell(row=i, column=col).value)
		
	# Opening about 20 tabs at the same time will cause mangadex to soft ban 
	# your ip. Hence time.sleep is used to keep a gap between tabs.
	for i, link in enumerate(links):
		if i % 15 == 0 and i != 0:
			time.sleep(10)
		print("Opening", link)
		webbrowser.open(link)

	print("Task Complete.")

if __name__ == "__main__":	
	if os.path.exists(folder_name):
		files = []
		for filename in os.listdir(folder_name):
			if filename.endswith(".xlsx"):
				files.append(filename)		
		if not files:
			print(msg)
		else:
			files.sort()
			# Convert to an indexed dictionary to help user choose a file more
			# easily.
			files = {str(i): filename for i, filename in enumerate(files, 1)}
			print("These files are available: \n")
			for i, filename in files.items():
				print("{}) {}".format(i, filename))			
			
			while True:	
				num = input("\nPlease enter index of file you wish to open: ")
				if num not in files.keys():
					 print("Please enter a valid number.")
				else:
					break
			print("\nOpening %s" % files[num])
			file_path = os.path.join(folder_name, files[num])
			link_opener(file_path)
	else:
		print(msg)


