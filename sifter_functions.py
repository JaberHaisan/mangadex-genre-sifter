# sifter_functions.py - Contains Manga class and most of the functions required for 
# mangadex_sifter to run.

import requests
import bs4
import threading
import re
import openpyxl
import os
import datetime

def format_link(link):
	"""Checks if link is of acceptable format for e.g.
	"https://mangadex.org/genre/5/comedy/0/2/" or
	"https://mangadex.org/genre/5/comedy". If it is, then returns 
	the link in the required format and the genre name.
	"""
	link_regex = re.compile(r'https://mangadex.org/genre/\d+/(\w+)')
	res = link_regex.search(link)
	assert res != None, "Link is not in an acceptable format."
	new_link = res.group(0) + "/0/"
	genre = res.group(1)
	return new_link, genre
	
	
def return_soup(url):
	"""Returns the soup object of the url."""
	res = requests.get(url)
	soup = bs4.BeautifulSoup(res.text, 'lxml')
	return soup
	
	
class Manga():
	
	def __init__(self, url):
		"""Initializes Manga class. Url should be from a manga in mangadex."""
		self.soup = return_soup(url)
		
	def rating(self):
		"""Finds the rating of manga object."""
		rating_elem = self.soup.select("span[class=text-primary]")
		try:
			return float(rating_elem[0].getText())
		except IndexError:
			print("There was an error in parsing this manga.")
			return 0.0	
			
	def genres(self):
		"""Returns genres of manga."""		
		genres = [elem.getText() for elem in self.soup.select(r'a[class="badge badge-secondary"]')]
		return genres
		
	def nationality(self):
		"""Returns nationality of manga."""
		country_name = self.soup.find_all("span", {"class": re.compile("rounded flag flag")})[0].get("title")
		return country_name		
		
	def has_rejected_genre(self, rejected_genres):
		"""Returns True if manga has a rejected genre."""
		rejected_genres = set(rejected_genres)
		return any(True if genre in rejected_genres else False for genre in self.genres())
	
	def is_rating_higher(self, min_rating):
		"""Returns True if rating is higher than min_rating."""
		return self.rating() > min_rating
		
		
def titles_finder(base_list_url, a_set):
	"""Updates given set with all manga titles from link."""
	titles = set()
	
	# Starts checking from base url and continues until there are no more
	# titles left.
	for i in range(1,1000):
		url = base_list_url + str(i)
		res = requests.get(url)
		soup = bs4.BeautifulSoup(res.text, "lxml")
		manga_elems = soup.select(r'a[class="ml-1 manga_title text-truncate"]')
		
		if not manga_elems:
			break
			
		for elem in manga_elems:
			titles.add(elem.getText())

	a_set.update(titles)
	
	
def start_and_join(threads):
	"""Starts and joins all given threads."""
	for thread in threads:
		thread.start()
	for thread in threads:
		thread.join()
			
				
def checked_set(user_id):
	"""Returns set of manga titles in user's manga lists."""
	base_link = "https://mangadex.org/list/%s/" % user_id
	checked = set()
	
	threads = []
	for i in range(1, 6 + 1):
		url = base_link + str(i) + "/2/"
		thread = threading.Thread(target=titles_finder, args=((url,checked)))
		threads.append(thread)
	start_and_join(threads)
		
	return checked


def manga_elem_finder(link, a_list):
	"""Finds all manga elems in given link and adds to a_list."""
	res = requests.get(link)
	soup = bs4.BeautifulSoup(res.text, "lxml")
	elems = soup.select(r'a[class="ml-1 manga_title text-truncate"]')
	
	if not elems:
		return
		
	a_list.extend(elems)


def all_manga_elem(base_link, max_page):
	"""Returns list with all manga elements starting from link up to given page."""
	manga_elems = []
	
	threads = []
	for i in range(1, max_page + 1):
		link = base_link + str(i)
		thread = threading.Thread(target=manga_elem_finder, args=(link, manga_elems))
		threads.append(thread)
	start_and_join(threads)
	
	return manga_elems
	
		
def save_mangas(mangas, genre, folder_name="Output"):
	"""Saves the mangas in an excel document in a folder with the given name."""
	wb = openpyxl.Workbook()
	ws = wb.active
		
	ws["A1"] = "Title"
	ws["B1"] = "Rating"
	ws["C1"] = "Link"
	ws.freeze_panes = "A2"

	ws.column_dimensions["A"].width = 70
	ws.column_dimensions["C"].width = 20

	r = 2
	for manga_name in mangas.keys():
		ws.cell(row=r, column=1).value = manga_name
		ws.cell(row=r, column=2).value = mangas[manga_name]['rating']
		ws.cell(row=r, column=3).value = mangas[manga_name]['link']
		r += 1
	
	os.makedirs(folder_name, exist_ok=True)
	
	current_date = datetime.datetime.now()
	date = datetime.datetime.strftime(current_date, "%y-%m-%d")
	
	file_name = date + " " + genre + ".xlsx"
	file_path = os.path.join(folder_name, file_name)
	wb.save(file_path)
